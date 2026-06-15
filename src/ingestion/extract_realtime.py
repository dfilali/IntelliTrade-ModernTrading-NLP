import os
import sys
import time
import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
from lxml import html
from datetime import datetime

# Ajout du chemin du projet au PYTHONPATH si exécuté en tant que script direct
base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
if base_dir not in sys.path:
    sys.path.append(base_dir)

import pandas_gbq
from src.utils.db import get_project_id
from src.utils.email_sender import send_market_report

def scrape_boursorama(company):
    symbol = company["Symbol"].split(".")[0]
    url = f'https://www.boursorama.com/cours/1rP{symbol}/'
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=15)
        if response.status_code == 200:
            root = html.fromstring(response.content)
            
            name = company["Name"]
            
            # Récupération du cours de l'entreprise
            element = root.xpath("/html/body/main/div/section/header/div/div/div[1]/div[1]/div/div[1]/span[1]")
            open_price = element[0].text_content().replace("  ", " ").strip() if element else None
            
            # XPaths pour High, Low, Volume, Close
            xpath_high = "/html/body/main/div/section/header/div/div/div[3]/div[1]/div/ul/li[3]/p[2]/span"
            xpath_low = "/html/body/main/div/section/header/div/div/div[3]/div[1]/div/ul/li[4]/p[2]/span"
            xpath_volume = "/html/body/main/div/section/header/div/div/div[3]/div[2]/div/ul/li[1]/p[2]/span"
            xpath_close = "/html/body/main/div/section/header/div/div/div[3]/div[1]/div/ul/li[2]/p[2]/span"
            
            high = root.xpath(xpath_high)
            low = root.xpath(xpath_low)
            volume = root.xpath(xpath_volume)
            close = root.xpath(xpath_close)
            
            high = high[0].text_content().replace("  ", " ").strip() if high else None
            low = low[0].text_content().replace("  ", " ").strip() if low else None
            volume = volume[0].text_content().replace("  ", " ").strip() if volume else None
            close = close[0].text_content().replace("  ", " ").strip() if close else None
            
            now = datetime.now()
            date_time = now.strftime("%Y-%m-%d %H:%M:%S")
            
            return {
                "Name": name, 
                "Open": open_price, 
                "High": high, 
                "Low": low, 
                "Close": close,
                "Volume": volume,
                "Date et Heure": date_time
            }
        else:
            print(f"Échec de la requête pour {company['Name']}. Code d'état: {response.status_code}")
            return None
    except Exception as e:
        print(f"Erreur lors du scraping de {company['Name']}: {e}")
        return None

def scrape_all_companies(companies):
    data = []
    for index, row in companies.iterrows():
        print(f"Scraping de {row['Name']} ({row['Symbol']})...")
        result = scrape_boursorama(row)
        if result:
            data.append(result)
        time.sleep(0.5)  # Pause pour éviter d'être banni par Boursorama
    return data

def insert_data_into_database(data):
    if not data:
        return
        
    try:
        # Remplacer "N/A" par None dans toutes les colonnes numériques
        cleaned_data = []
        for entry in data:
            item = entry.copy()
            for key, value in item.items():
                if key not in ['Name', 'Date et Heure']:
                    if pd.isna(value) or value == "N/A" or value is None:
                        item[key] = None
                    else:
                        val_str = str(value).replace(' ', '').replace('\xa0', '').replace(',', '.')
                        try:
                            item[key] = float(val_str)
                        except ValueError:
                            item[key] = None
            cleaned_data.append(item)

        df = pd.DataFrame(cleaned_data)
        
        # Insérer les données dans la table Google BigQuery brute
        project_id = get_project_id()
        nom_table = 'raw_trading.data_trading'
        print(f"Chargement des données temps réel dans BigQuery : '{nom_table}' (if_exists='append')...")
        pandas_gbq.to_gbq(df, nom_table, project_id=project_id, if_exists='append')
        print(f"Données en temps réel ajoutées avec succès à '{nom_table}' dans BigQuery.")
    except Exception as e:
        print(f"Erreur lors de l'insertion dans BigQuery : {e}")

def save_to_excel(data):
    if not data:
        return None
        
    try:
        output_dir = os.path.join(base_dir, 'data', 'output')
        os.makedirs(output_dir, exist_ok=True)
        filename = os.path.join(output_dir, f'donnees_boursorama_{datetime.now().strftime("%Y%m%d")}.xlsx')
        
        # Vérifier si le fichier existe déjà
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            # Ajouter les en-têtes si le fichier est nouveau
            headers = list(data[0].keys())
            ws.append(headers)

        # Écrire les données dans le fichier Excel
        for row_data in data:
            ws.append(list(row_data.values()))

        # Sauvegarder le fichier Excel
        wb.save(filename)
        print(f"Données écrites avec succès dans le fichier Excel : {filename}")
        return filename
    except Exception as e:
        print(f"Erreur lors de l'enregistrement Excel : {e}")
        return None

def job():
    print("Exécution de la tâche de scraping en temps réel...")
    company_referal_path = os.path.join(base_dir, 'data', 'input', 'liste_cac40.xlsx')
    if not os.path.exists(company_referal_path):
        print(f"Fichier de référence introuvable: {company_referal_path}")
        return None
        
    cac40_companies = pd.read_excel(company_referal_path)
    all_data = scrape_all_companies(cac40_companies)
    if all_data:
        insert_data_into_database(all_data)
        excel_path = save_to_excel(all_data)
        return excel_path
    return None

def start_scheduled_loop():
    # Démarre la boucle de scraping de 9h à 18h
    # (Configuration identique à l'original)
    start_time = datetime.now().replace(hour=11, minute=15, second=0, microsecond=0)
    end_time = datetime.now().replace(hour=11, minute=17, second=0, microsecond=0)

    excel_file = None
    while datetime.now() < end_time:
        if datetime.now() >= start_time:
            excel_file = job()
        time.sleep(30)
        
    return excel_file

if __name__ == '__main__':
    current_day = datetime.now().weekday()
    # Si lundi (0) à vendredi (4)
    if 0 <= current_day <= 4:
        excel_file = start_scheduled_loop()
        
        # Envoi de mail en clôture
        if excel_file:
            send_market_report(excel_file)
    else:
        print("Le programme de scraping automatique ne s'exécute pas le week-end.")
